#!/usr/bin/env python3
"""
ingest.py — build wartears.db from wartears.org's daily xlsx snapshot.

Source: https://wartears.org publishes a daily snapshot of their full database
as a single xlsx at /static/wartears-opendata.xlsx — there's no streaming API
and no diff endpoint, so every fetch is the full ~65 MB workbook (regenerated
overnight via openpyxl on their end).

Five sheets, Russian-named (the only canonical form):
  Метаданные                 — single key/value: snapshot generation timestamp.
  Записи                     — RECORDS about persons (kind=1), organisations
                               (kind=2), or other (kind=3). Up to 8 tags, 8
                               free-form attributes, and 4 image URLs per row.
                               The canonical tags are "Погиб" (DEAD) and
                               "В плену" (captured).
  Теги                       — tag lookup (id → name).
  Отношения                  — oriented graph edges between records (e.g.
                               "serves-in", "is-part-of").
  Разновидности отношений    — relationship-kind lookup.

APPEND-ONLY / EDIT-VERSIONED for records (mirrors the other ingest scripts).
A stored row is never mutated or deleted; an upstream edit inserts a NEW row
keyed by our ingest timestamp `scraped_at`. We use wartears' own per-record
`updated_at` as the change signal — a fetched record whose `updated_at` matches
the latest stored version is skipped, so a daily re-fetch of an unchanged
record costs nothing.

Tags / attrs / images are stored in normalised side-tables versioned alongside
the record (same `(record_id, scraped_at)` key), so a record-version is a
self-contained snapshot. `record_tags` carries an index on `tag` so the common
"all DEAD persons" query is a single scan, not a JSON unbag.

Lookup tables (tags, relationship_kinds, relationships, source_meta) are
replaced wholesale each run — they're tiny and not historically interesting.

Single-transaction write with rollback-on-guard-failure: an absolute records
floor and a no-shrink check abort the build without touching the DB if the
fetch looks broken.

Requires openpyxl — xlsx is a zip of XML and the 246k-row records sheet needs
streaming to stay under memory. openpyxl's read-only iterator handles it in
~75s and constant memory.
"""

from __future__ import annotations

import argparse
import io
import os
import sqlite3
import sys
import time
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

UPSTREAM_URL = "https://wartears.org/static/wartears-opendata.xlsx"

SHEET_META = "Метаданные"
SHEET_RECORDS = "Записи"
SHEET_TAGS = "Теги"
SHEET_RELATIONSHIPS = "Отношения"
SHEET_REL_KINDS = "Разновидности отношений"

# Records floor: wartears tracked ~245k records by mid-2026. A fetch returning
# far less is a broken/truncated workbook — refuse to write.
MIN_RECORDS_FLOOR = 100_000

# Batch size for the streaming insert loop. 5000 keeps each executemany cheap
# while avoiding per-row overhead; the records sheet has ~246k rows.
BATCH = 5000

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DB_NAME = os.environ.get("WARTEARS_DB_NAME", "wartears.db")

SCHEMA = """
CREATE TABLE IF NOT EXISTS records (
    id INTEGER NOT NULL,
    scraped_at TEXT NOT NULL,
    kind INTEGER,
    name1 TEXT, name2 TEXT, name3 TEXT, name4 TEXT,
    birth_year INTEGER,
    public_info TEXT,
    updated_at TEXT,
    PRIMARY KEY (id, scraped_at)
);
CREATE INDEX IF NOT EXISTS records_id ON records(id);

CREATE TABLE IF NOT EXISTS record_tags (
    record_id INTEGER NOT NULL,
    scraped_at TEXT NOT NULL,
    tag TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS record_tags_rec ON record_tags(record_id, scraped_at);
CREATE INDEX IF NOT EXISTS record_tags_tag ON record_tags(tag);

CREATE TABLE IF NOT EXISTS record_attrs (
    record_id INTEGER NOT NULL,
    scraped_at TEXT NOT NULL,
    name TEXT,
    value TEXT
);
CREATE INDEX IF NOT EXISTS record_attrs_rec ON record_attrs(record_id, scraped_at);

CREATE TABLE IF NOT EXISTS record_images (
    record_id INTEGER NOT NULL,
    scraped_at TEXT NOT NULL,
    url TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS record_images_rec ON record_images(record_id, scraped_at);

CREATE TABLE IF NOT EXISTS tags (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS relationships (
    id INTEGER PRIMARY KEY,
    kind INTEGER,
    left_id INTEGER,
    right_id INTEGER
);
CREATE INDEX IF NOT EXISTS relationships_left ON relationships(left_id);
CREATE INDEX IF NOT EXISTS relationships_right ON relationships(right_id);

CREATE TABLE IF NOT EXISTS relationship_kinds (
    id INTEGER PRIMARY KEY,
    code TEXT,
    left_kind INTEGER,
    right_kind INTEGER,
    name_lr TEXT,
    name_rl TEXT,
    symmetric INTEGER
);

CREATE TABLE IF NOT EXISTS source_meta (
    key TEXT PRIMARY KEY,
    value TEXT
);
"""


def fetch_xlsx() -> io.BytesIO:
    """Download the daily snapshot into memory. Single ~65 MB transfer.

    openpyxl accepts a file-like object, so we don't write to disk — keeps
    the script consistent with ru_losses / missile_attacks (fetch → parse,
    no on-disk cache; pass `--in` for a local file during dev).
    """
    req = urllib.request.Request(UPSTREAM_URL, headers={"User-Agent": "sbs-stats-ingest"})
    print(f"[fetch] {UPSTREAM_URL}")
    t0 = time.time()
    buf = io.BytesIO()
    with urllib.request.urlopen(req, timeout=300) as resp:
        if resp.status != 200:
            raise RuntimeError(f"{UPSTREAM_URL} returned HTTP {resp.status}")
        while True:
            chunk = resp.read(1 << 20)
            if not chunk:
                break
            buf.write(chunk)
    print(f"[fetch] {buf.tell():,} bytes in {time.time() - t0:.1f}s")
    buf.seek(0)
    return buf


def _coerce_date(v) -> str | None:
    """Date-only ISO string for an upstream datetime cell (or None)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)


def _iter_records(ws):
    """Stream records from the Записи sheet as dicts."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    h = {name: i for i, name in enumerate(header) if name is not None}
    tag_cols = [h[k] for k in header if isinstance(k, str) and k.startswith("Тег")]
    attr_n_cols = [h[k] for k in header if isinstance(k, str) and k.startswith("Название атрибута")]
    attr_v_cols = [h[k] for k in header if isinstance(k, str) and k.startswith("Значение атрибута")]
    img_cols = [h[k] for k in header if isinstance(k, str) and k.startswith("Изображение")]

    for r in rows:
        if r[h["ID"]] is None:
            continue
        yield {
            "id": int(r[h["ID"]]),
            "kind": r[h["Разновидность"]],
            "name1": r[h["ФИО 1"]],
            "name2": r[h["ФИО 2"]],
            "name3": r[h["ФИО 3"]],
            "name4": r[h["ФИО 4"]],
            "birth_year": r[h["Год рождения"]],
            "public_info": r[h["Публичная информация"]],
            "updated_at": _coerce_date(r[h["Дата обновления"]]),
            "tags": [r[c] for c in tag_cols if r[c]],
            "attrs": [(r[n], r[v]) for n, v in zip(attr_n_cols, attr_v_cols) if r[n] or r[v]],
            "images": [r[c] for c in img_cols if r[c]],
        }


def _all_rows(ws):
    """Return [(cells)] for every non-header data row of a small sheet."""
    rows = ws.iter_rows(values_only=True)
    next(rows)  # discard header
    return [r for r in rows if r and r[0] is not None]


def build(db_path: Path, xlsx_source) -> dict:
    """Append-version `records` from xlsx_source into db_path. Replace lookup tables.

    `xlsx_source` is either a Path or an in-memory file-like (BytesIO).
    Returns a summary dict. Aborts (raises) without writing if the fetch fails
    the row-count floor or would shrink an existing dataset.
    """
    import openpyxl
    print(f"[parse] opening workbook")
    wb = openpyxl.load_workbook(xlsx_source, read_only=True, data_only=True)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(SCHEMA)

        # Latest stored updated_at per record_id — used to skip unchanged rows.
        stored = dict(conn.execute(
            "SELECT id, updated_at FROM ("
            " SELECT id, updated_at, "
            "        ROW_NUMBER() OVER (PARTITION BY id ORDER BY scraped_at DESC) AS rn"
            " FROM records"
            ") WHERE rn = 1"
        ).fetchall())
        print(f"[change] {len(stored):,} records already in DB")

        # Refresh small lookup tables. Inside the same transaction as the
        # records inserts, so a guard-triggered rollback restores them too.
        conn.execute("DELETE FROM tags")
        tags = [(int(r[0]), r[1]) for r in _all_rows(wb[SHEET_TAGS])]
        conn.executemany("INSERT INTO tags (id, name) VALUES (?, ?)", tags)

        conn.execute("DELETE FROM relationship_kinds")
        rks = [
            (int(r[0]), r[1], r[2], r[3], r[4], r[5], 1 if r[6] else 0)
            for r in _all_rows(wb[SHEET_REL_KINDS])
        ]
        conn.executemany(
            "INSERT INTO relationship_kinds "
            "(id, code, left_kind, right_kind, name_lr, name_rl, symmetric) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)", rks,
        )

        conn.execute("DELETE FROM relationships")
        rels = [
            (int(r[0]), r[1],
             int(r[2]) if r[2] is not None else None,
             int(r[3]) if r[3] is not None else None)
            for r in _all_rows(wb[SHEET_RELATIONSHIPS])
        ]
        conn.executemany(
            "INSERT INTO relationships (id, kind, left_id, right_id) VALUES (?, ?, ?, ?)",
            rels,
        )

        conn.execute("DELETE FROM source_meta")
        meta = [(r[0], _coerce_date(r[1]) if isinstance(r[1], datetime) else
                 (str(r[1]) if r[1] is not None else None))
                for r in _all_rows(wb[SHEET_META])]
        conn.executemany("INSERT INTO source_meta (key, value) VALUES (?, ?)", meta)

        print(f"[lookup] tags={len(tags)} rel_kinds={len(rks)} relationships={len(rels):,} meta={len(meta)}")
        for k, v in meta:
            print(f"[lookup] meta: {k}={v}")

        # Stream the records sheet — the only big one. Skip rows whose
        # upstream updated_at matches what we already stored.
        scraped_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
        rec_b, tag_b, attr_b, img_b = [], [], [], []
        total = inserted = unchanged = 0
        t0 = time.time()

        def flush():
            if rec_b:
                conn.executemany(
                    "INSERT INTO records (id, scraped_at, kind, name1, name2, name3, name4, "
                    "birth_year, public_info, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)", rec_b,
                )
            if tag_b:
                conn.executemany(
                    "INSERT INTO record_tags (record_id, scraped_at, tag) VALUES (?,?,?)", tag_b,
                )
            if attr_b:
                conn.executemany(
                    "INSERT INTO record_attrs (record_id, scraped_at, name, value) VALUES (?,?,?,?)", attr_b,
                )
            if img_b:
                conn.executemany(
                    "INSERT INTO record_images (record_id, scraped_at, url) VALUES (?,?,?)", img_b,
                )
            rec_b.clear(); tag_b.clear(); attr_b.clear(); img_b.clear()

        for rec in _iter_records(wb[SHEET_RECORDS]):
            total += 1
            # `updated_at is None` records (rare) re-insert every run — we
            # can't tell whether they changed without a full-row compare.
            if rec["updated_at"] is not None and stored.get(rec["id"]) == rec["updated_at"]:
                unchanged += 1
                continue
            inserted += 1
            rec_b.append((
                rec["id"], scraped_at, rec["kind"],
                rec["name1"], rec["name2"], rec["name3"], rec["name4"],
                rec["birth_year"], rec["public_info"], rec["updated_at"],
            ))
            for t in rec["tags"]:
                tag_b.append((rec["id"], scraped_at, t))
            for n, v in rec["attrs"]:
                attr_b.append((rec["id"], scraped_at, n, v))
            for url in rec["images"]:
                img_b.append((rec["id"], scraped_at, url))
            if len(rec_b) >= BATCH:
                flush()
        flush()
        print(f"[parse] {total:,} records in {time.time() - t0:.1f}s "
              f"({inserted:,} inserted, {unchanged:,} unchanged)")

        # Guard 1: absolute floor.
        if total < MIN_RECORDS_FLOOR:
            raise RuntimeError(
                f"parsed only {total} records (< floor {MIN_RECORDS_FLOOR}) — "
                f"refusing to write {db_path}; upstream likely returned a "
                f"truncated workbook."
            )
        # Guard 2: shrink — total fetched < distinct ids already stored.
        if total < len(stored):
            raise RuntimeError(
                f"parsed {total} records but DB already has {len(stored)} distinct "
                f"records — refusing to write a shrinking dataset into {db_path}."
            )

        conn.commit()
        # VACUUM after the big append to reclaim space from any prior runs.
        conn.execute("VACUUM")

        distinct = conn.execute("SELECT COUNT(DISTINCT id) FROM records").fetchone()[0]
        versions = conn.execute("SELECT COUNT(*) FROM records").fetchone()[0]
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "fetched": total, "inserted": inserted, "unchanged": unchanged,
        "distinct_records": distinct, "record_versions": versions,
        "tags": len(tags), "relationships": len(rels), "rel_kinds": len(rks),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Build wartears.db from the daily xlsx snapshot")
    ap.add_argument(
        "--in", dest="in_path", default=os.environ.get("WARTEARS_XLSX_PATH"),
        help="path to an already-downloaded wartears-opendata.xlsx (default: fetch from upstream)",
    )
    ap.add_argument(
        "--out", default=os.environ.get(
            "WARTEARS_DB_PATH", str(SCRIPT_DIR / "output" / DEFAULT_DB_NAME)),
        help="output SQLite path (default: scripts/wartears/output/%s)" % DEFAULT_DB_NAME,
    )
    args = ap.parse_args()

    if args.in_path:
        xlsx_source = Path(args.in_path)
        if not xlsx_source.exists():
            print(f"--in path does not exist: {xlsx_source}", file=sys.stderr)
            return 2
    else:
        xlsx_source = fetch_xlsx()

    out = Path(args.out)
    s = build(out, xlsx_source)
    print(
        f"==> {s['fetched']:,} fetched, {s['inserted']:,} inserted, "
        f"{s['unchanged']:,} unchanged | {s['distinct_records']:,} distinct records "
        f"({s['record_versions']:,} versions); tags={s['tags']}, "
        f"relationships={s['relationships']:,}, rel_kinds={s['rel_kinds']}\n"
        f"==> {out} ({out.stat().st_size:,} bytes)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
