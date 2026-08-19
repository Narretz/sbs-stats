#!/usr/bin/env python3
"""
ingest.py — build ua-losses.db from ualosses.org's UKR_ualosses_Personnel.xlsx.

Source: ualosses.org publishes a workbook of confirmed Ukrainian military
personnel losses — individually researched, named records — on Kaggle
(`ol4ubert/confirmed-ukrainian-military-personnel-losses`), re-uploaded
periodically so Kaggle keeps a numbered version history.

We build a per-day series from two sheets:

  ByDay        — col 1 is the day's total losses recorded on that date.
  ByDayStatus  — Dead / Missing / Prisoner / Released split of that day.

**Parsed positionally, not by header name**, because ualosses has renamed the
columns repeatedly across versions (ByDay col 1 has been `NumberDay`,
`NumberHelp`, and `Number`; the status sheet gained `Released` only later, and
didn't exist at all before ~v12 when the project tracked deaths only). Positions
have stayed put; names haven't. When `ByDayStatus` is absent (early
deaths-only versions), every loss is a death, so `dead = number`.

`released` = a POW since freed (still dated at the original capture event).

APPEND-ONLY / EDIT-VERSIONED (mirrors mediazona / ru_losses / wartears). A stored
row is never mutated or deleted. Each row is one *version* of that date's numbers,
tagged with `scraped_at`. ualosses continuously revises past days — both by adding
newly-identified people and by **reclassifying** existing ones (missing → dead /
POW), which shifts a date's split without changing its `number`. Each run compares
parsed values against the latest stored version per date and INSERTs a NEW row
only where they differ. Replaying the Kaggle version history oldest→newest (see
backfill.sh) reconstructs that revision/reclassification history retroactively.

A row-count floor + a no-shrink guard abort the build (without writing) if the
workbook looks truncated.

Requires openpyxl. Kaggle download (--version) needs KAGGLE_USERNAME / KAGGLE_KEY.
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import date as date_cls, datetime, timezone
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
DEFAULT_DB_NAME = os.environ.get("UA_LOSSES_DB_NAME", "ua-losses.db")
DEFAULT_XLSX = os.environ.get(
    "UA_LOSSES_XLSX", str(REPO_ROOT / "sources" / "ualosses" / "UKR_ualosses_Personnel.xlsx")
)
KAGGLE_REF_DEFAULT = os.environ.get(
    "UA_LOSSES_KAGGLE_REF", "ol4ubert/confirmed-ukrainian-military-personnel-losses"
)

SHEET_BYDAY = "ByDay"
SHEET_STATUS = "ByDayStatus"

# Integer metric columns, in storage order. `number` is the day's total losses;
# the status split follows. (Demographic columns like male / mean age are not
# stored — their position/name churns across versions and they aren't charted.)
STATUS_COLS = ["dead", "missing", "prisoner", "released"]
METRIC_INT_COLS = ["number", *STATUS_COLS]

# The war began 2022-02-24; ualosses' workbook also carries a long pre-war tail
# (Donbas-era and older records) irrelevant to this dashboard, so we default to
# war-scoping the DB. Override with --all-dates.
WAR_START = "2022-02-24"

# ~1.5 years of war by any real export. Far below this = truncated workbook.
MIN_ROWS_FLOOR = 365


# ── value coercion ────────────────────────────────────────────────────────────
def _int0(v) -> int:
    """A blank/None cell in a daily count means a genuine zero for that day."""
    if v is None or v == "":
        return 0
    return int(v)


def _date_iso(v) -> str | None:
    """Normalise a Date cell (datetime, date, or 'YYYY-MM-DD…' string) to
    'YYYY-MM-DD'. Returns None for a blank/unparseable cell (e.g. header echo)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date_cls):
        return v.isoformat()
    head = str(v).strip()[:10]
    try:
        date_cls.fromisoformat(head)
        return head
    except ValueError:
        return None


# ── sheet parsing (positional; pure, so tests need no xlsx) ────────────────────
def index_byday(sheet_rows) -> dict[str, int]:
    """{date_iso: daily total}. Column 0 = Date, column 1 = the day's count."""
    it = iter(sheet_rows)
    next(it, None)  # header
    out: dict[str, int] = {}
    for r in it:
        d = _date_iso(r[0] if r else None)
        if d is None:
            continue
        out[d] = _int0(r[1]) if len(r) > 1 else 0
    return out


def index_status(sheet_rows) -> dict[str, tuple[int, int, int, int]]:
    """{date_iso: (dead, missing, prisoner, released)}. Columns: Date, Dead,
    Missing, Prisoner, [Released]. `Released` only appears in later versions —
    default 0 when the column is absent."""
    it = iter(sheet_rows)
    next(it, None)  # header
    out: dict[str, tuple[int, int, int, int]] = {}
    for r in it:
        d = _date_iso(r[0] if r else None)
        if d is None:
            continue
        out[d] = (
            _int0(r[1]) if len(r) > 1 else 0,
            _int0(r[2]) if len(r) > 2 else 0,
            _int0(r[3]) if len(r) > 3 else 0,
            _int0(r[4]) if len(r) > 4 else 0,
        )
    return out


def join_daily(byday: dict, status: dict, since: str | None) -> list[tuple]:
    """Join the daily total with its status split into storage tuples:
        (date, number, dead, missing, prisoner, released)
    A date with no status row (deaths-only versions, which have no ByDayStatus
    sheet at all) is treated as all-dead. Rows before `since` are dropped."""
    rows: list[tuple] = []
    for d in sorted(byday):
        if since and d < since:
            continue
        number = byday[d]
        if d in status:
            dead, missing, prisoner, released = status[d]
        else:
            dead, missing, prisoner, released = number, 0, 0, 0
        rows.append((d, number, dead, missing, prisoner, released))
    return rows


def parse_workbook(xlsx_path: Path, since: str | None) -> list[tuple]:
    """Read ByDay (+ ByDayStatus when present) and return storage tuples."""
    import openpyxl  # lazy — keeps `--help` cheap without the dep

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SHEET_BYDAY not in wb.sheetnames:
            raise RuntimeError(
                f"{xlsx_path} has no '{SHEET_BYDAY}' sheet; found {wb.sheetnames}. "
                f"Is this the ualosses UKR_ualosses_Personnel workbook?"
            )
        byday = index_byday(wb[SHEET_BYDAY].iter_rows(values_only=True))
        status = (index_status(wb[SHEET_STATUS].iter_rows(values_only=True))
                  if SHEET_STATUS in wb.sheetnames else {})
    finally:
        wb.close()
    return join_daily(byday, status, since)


# ── Kaggle version download ───────────────────────────────────────────────────
def latest_kaggle_version(ref: str) -> int:
    """Return a Kaggle dataset's current (max) version number via the API."""
    import base64
    import json
    import urllib.request

    user = os.environ.get("KAGGLE_USERNAME")
    key = os.environ.get("KAGGLE_KEY")
    if not user or not key:
        raise SystemExit(
            "KAGGLE_USERNAME / KAGGLE_KEY must be set for --latest "
            "(e.g. `set -a; . ./.env.kaggle; set +a`)."
        )
    tok = base64.b64encode(f"{user}:{key}".encode()).decode()
    req = urllib.request.Request(
        f"https://www.kaggle.com/api/v1/datasets/view/{ref}",
        headers={"Authorization": f"Basic {tok}"})
    with urllib.request.urlopen(req, timeout=60) as r:
        data = json.load(r)
    n = data.get("currentVersionNumber")
    if not isinstance(n, int):
        raise RuntimeError(f"no currentVersionNumber in Kaggle view for {ref}: {data!r}")
    return n


def download_kaggle_version(ref: str, version: int, dest_dir: Path) -> Path:
    """Download one Kaggle dataset version's xlsx via the public API (HTTP basic
    auth with KAGGLE_USERNAME / KAGGLE_KEY) and return the extracted path."""
    import base64
    import io
    import urllib.request
    import zipfile

    user = os.environ.get("KAGGLE_USERNAME")
    key = os.environ.get("KAGGLE_KEY")
    if not user or not key:
        raise SystemExit(
            "KAGGLE_USERNAME / KAGGLE_KEY must be set for --version "
            "(e.g. `set -a; . ./.env.kaggle; set +a`)."
        )
    url = (f"https://www.kaggle.com/api/v1/datasets/download/{ref}"
           f"?datasetVersionNumber={version}")
    tok = base64.b64encode(f"{user}:{key}".encode()).decode()
    req = urllib.request.Request(url, headers={"Authorization": f"Basic {tok}"})
    print(f"[kaggle] downloading {ref} v{version} …", file=sys.stderr)
    with urllib.request.urlopen(req, timeout=600) as r:
        blob = r.read()

    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not names:
            raise RuntimeError(
                f"no .xlsx in Kaggle download for {ref} v{version}: {zf.namelist()}")
        zf.extract(names[0], dest_dir)
        return dest_dir / names[0]


# ── DB build (append-version, guarded) ────────────────────────────────────────
def build(db_path: Path, rows: list[tuple], scraped_at: str) -> dict:
    """Append changed/new date-versions into db_path. Never mutates/deletes rows.

    Returns a summary dict: new / revised / unchanged / distinct. Aborts (raises)
    without writing if the payload looks broken or would shrink the dataset.
    """
    # Guard 1: absolute floor.
    if len(rows) < MIN_ROWS_FLOOR:
        raise RuntimeError(
            f"parsed only {len(rows)} day-rows (< floor {MIN_ROWS_FLOOR}) — refusing "
            f"to write {db_path}; the workbook is probably truncated."
        )

    int_cols_sql = ", ".join(f"{c} INTEGER NOT NULL DEFAULT 0" for c in METRIC_INT_COLS)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(
            f"""
            CREATE TABLE IF NOT EXISTS daily_losses (
                date TEXT NOT NULL,
                scraped_at TEXT NOT NULL,
                {int_cols_sql},
                PRIMARY KEY (date, scraped_at)
            );
            """
        )

        # Latest stored version per date, for change detection (mirrors the
        # ru_losses / mediazona LATEST_PER_DATE pattern).
        metric_sql = ", ".join(METRIC_INT_COLS)
        latest = {
            r[0]: tuple(r[1:])
            for r in conn.execute(
                f"""
                SELECT d.date, {metric_sql}
                FROM daily_losses d
                JOIN (SELECT date, MAX(scraped_at) AS ms FROM daily_losses GROUP BY date) l
                  ON d.date = l.date AND d.scraped_at = l.ms
                """
            ).fetchall()
        }

        # Guard 2: shrink — a fresh export should never carry fewer distinct
        # dates than what's already stored. Catches a half-empty workbook that
        # slipped past Guard 1.
        if len(rows) < len(latest):
            raise RuntimeError(
                f"parsed {len(rows)} dates but DB already has {len(latest)} — "
                f"refusing to write a shrinking dataset into {db_path}."
            )

        to_insert: list[tuple] = []
        new = revised = unchanged = 0
        for row in rows:
            d, *vals = row
            prior = latest.get(d)
            if prior is None:
                new += 1
                to_insert.append((d, scraped_at, *vals))
            elif prior != tuple(vals):
                revised += 1
                to_insert.append((d, scraped_at, *vals))
            else:
                unchanged += 1

        if to_insert:
            cols = ", ".join(["date", "scraped_at"] + METRIC_INT_COLS)
            ph = ", ".join(["?"] * (2 + len(METRIC_INT_COLS)))
            conn.executemany(f"INSERT INTO daily_losses ({cols}) VALUES ({ph})", to_insert)
            conn.commit()
            conn.execute("VACUUM")
            conn.commit()

        distinct = conn.execute("SELECT COUNT(DISTINCT date) FROM daily_losses").fetchone()[0]
    finally:
        conn.close()
    return {"new": new, "revised": revised, "unchanged": unchanged, "distinct": distinct}


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Append-version ua-losses.db from ualosses.org's personnel workbook.",
    )
    ap.add_argument("--xlsx", default=DEFAULT_XLSX,
                    help="local UKR_ualosses_Personnel.xlsx (default: %(default)s)")
    ap.add_argument("--version", type=int, default=None, metavar="N",
                    help="download this Kaggle dataset version instead of --xlsx")
    ap.add_argument("--latest", action="store_true",
                    help="download the current (latest) Kaggle version — for CI")
    ap.add_argument("--kaggle-ref", default=KAGGLE_REF_DEFAULT,
                    help="Kaggle dataset ref for --version/--latest (default: %(default)s)")
    ap.add_argument("--since", default=WAR_START,
                    help="drop dates before this YYYY-MM-DD (default: %(default)s)")
    ap.add_argument("--all-dates", action="store_true",
                    help="keep the full pre-war tail (overrides --since)")
    ap.add_argument("--as-of", default=None, metavar="YYYY-MM-DD",
                    help="stamp scraped_at with this date. Default: the data's own "
                         "vintage (max date) when --version is used, else now — so a "
                         "backfilled version lands chronologically, not all at 'now'.")
    ap.add_argument("--out", default=os.environ.get(
        "UA_LOSSES_DB_PATH", str(SCRIPT_DIR / "output" / DEFAULT_DB_NAME)),
        help="output SQLite path (default: scripts/ua_losses/output/%s)" % DEFAULT_DB_NAME)
    ap.add_argument("--dry-run", action="store_true",
                    help="parse + report what would change, without writing")
    args = ap.parse_args()

    since = None if args.all_dates else args.since
    if since is not None:
        try:
            date_cls.fromisoformat(since)
        except ValueError:
            raise SystemExit(f"--since must be YYYY-MM-DD, got {since!r}")
    if args.as_of is not None:
        try:
            date_cls.fromisoformat(args.as_of)
        except ValueError:
            raise SystemExit(f"--as-of must be YYYY-MM-DD, got {args.as_of!r}")

    if args.latest:
        args.version = latest_kaggle_version(args.kaggle_ref)
        print(f"[kaggle] latest version = v{args.version}", file=sys.stderr)

    if args.version is not None:
        xlsx = download_kaggle_version(
            args.kaggle_ref, args.version, SCRIPT_DIR / "output" / f"v{args.version}")
    else:
        xlsx = Path(args.xlsx)
        if not xlsx.exists():
            raise SystemExit(f"xlsx not found: {xlsx}")
    rows = parse_workbook(xlsx, since)

    if not rows:
        raise SystemExit("no dated rows parsed — nothing to write.")
    vintage = rows[-1][0]  # rows are date-sorted; last is the newest event date
    total = sum(r[1] for r in rows)
    print(f"[parse] {len(rows)} day-rows {rows[0][0]}..{vintage}; sum(number)={total}"
          f"{'' if since is None else f'  (since {since})'}")

    # scraped_at: explicit --as-of wins; else the version's vintage for backfill
    # (distinct + chronological per version); else now for a live local run.
    if args.as_of is not None:
        scraped_at = f"{args.as_of}T00:00:00+00:00"
    elif args.version is not None:
        scraped_at = f"{vintage}T00:00:00+00:00"
    else:
        scraped_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    if args.dry_run:
        out = Path(args.out)
        prior_n = 0
        if out.exists():
            conn = sqlite3.connect(out)
            try:
                prior_n = conn.execute("SELECT COUNT(DISTINCT date) FROM daily_losses").fetchone()[0]
            except sqlite3.OperationalError:
                prior_n = 0
            finally:
                conn.close()
        print(f"[dry-run] scraped_at={scraped_at}; floor={MIN_ROWS_FLOOR}, "
              f"existing distinct dates={prior_n}; no write.")
        return 0

    out = Path(args.out)
    s = build(out, rows, scraped_at)
    print(
        f"==> {s['new']:>4} new, {s['revised']:>4} revised, {s['unchanged']:>4} unchanged "
        f"-> {s['new'] + s['revised']} inserted; {s['distinct']} distinct dates\n"
        f"==> scraped_at={scraped_at}\n"
        f"==> {out} ({out.stat().st_size} bytes)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
